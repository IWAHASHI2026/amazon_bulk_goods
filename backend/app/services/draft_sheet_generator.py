"""
Generate a draft Excel sheet for bulk product registration.
The sheet is dynamically built from the active template's column_schema and dropdown_schema.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO
from datetime import datetime


# Fields shown in the "基本情報" (Basic Info) vertical sheet.
# Each tuple: (display_name, technical_name_pattern, special_key)
# technical_name_pattern is used to find the matching column in column_schema.
# If special_key is set, the value maps to a Product field instead of parent_data.
BASIC_INFO_FIELDS = [
    ("商品名（管理用）", None, "_app_name"),
    ("SKU", "contribution_sku", None),
    ("ブランド名", "brand[", None),
    ("商品名（Amazon表示）", "item_name[", None),
    ("商品説明", "product_description[", None),
    ("仕様1", "bullet_point[", None),  # will match #1
    ("仕様2", "bullet_point[", None),  # will match #2
    ("仕様3", "bullet_point[", None),  # will match #3
    ("仕様4", "bullet_point[", None),  # will match #4
    ("仕様5", "bullet_point[", None),  # will match #5
    ("商品IDタイプ", "amzn1.volt.ca.product_id_type", None),
    ("商品ID", "amzn1.volt.ca.product_id_value", None),
    ("原産国", "country_of_origin[", None),
    ("バリエーションテーマ", "variation_theme", "_variation_theme"),
]

# Variation-specific field patterns for the "バリエーション" sheet
VARIATION_FIELD_PATTERNS = [
    ("子SKU", "contribution_sku"),
    ("色", "color["),
    ("サイズ", "size["),
    ("スタイル", "style["),
    ("素材", "material["),
    ("パターン", "pattern["),
    ("販売価格", "purchasable_offer["),
    ("B2B販売価格", "purchasable_offer["),  # second match
    ("在庫数", "fulfillment_availability"),
    ("メイン画像URL", "main_product_image_locator["),
    ("商品IDタイプ", "amzn1.volt.ca.product_id_type"),
    ("商品ID", "amzn1.volt.ca.product_id_value"),
]

# Styles
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LABEL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
LABEL_FONT = Font(bold=True, size=11)
SAMPLE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SAMPLE_FONT = Font(color="808080", italic=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _find_technical_name(column_schema: list[dict], pattern: str, occurrence: int = 1) -> str | None:
    """Find a technical_name in column_schema matching a pattern prefix."""
    count = 0
    for col in column_schema:
        tn = col["technical_name"]
        if tn.startswith(pattern) or tn == pattern:
            count += 1
            if count == occurrence:
                return tn
    return None


def _resolve_basic_info_fields(column_schema: list[dict]) -> list[tuple[str, str]]:
    """Resolve BASIC_INFO_FIELDS to (display_name, technical_name) pairs."""
    result = []
    bullet_count = 0
    for display, pattern, special_key in BASIC_INFO_FIELDS:
        if special_key:
            result.append((display, special_key))
            continue
        if pattern == "bullet_point[":
            bullet_count += 1
            tn = _find_technical_name(column_schema, pattern, bullet_count)
        else:
            tn = _find_technical_name(column_schema, pattern)
        if tn:
            result.append((display, tn))
        elif special_key is None and pattern:
            # Pattern didn't match - skip this field
            pass
    return result


def _resolve_variation_fields(column_schema: list[dict]) -> list[tuple[str, str]]:
    """Resolve VARIATION_FIELD_PATTERNS to (display_name, technical_name) pairs."""
    result = []
    seen = set()
    for display, pattern in VARIATION_FIELD_PATTERNS:
        for col in column_schema:
            tn = col["technical_name"]
            if tn not in seen and (tn.startswith(pattern) or tn == pattern):
                result.append((display, tn))
                seen.add(tn)
                break
    return result


def _add_dropdown_validation(ws, cell_range: str, values: list[str]):
    """Add a dropdown data validation if the values list fits Excel's limit."""
    formula = ",".join(values)
    if len(formula) > 255:
        # Excel data validation formula1 limit; skip overly long lists
        return
    dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
    dv.error = "リストから選択してください"
    dv.errorTitle = "入力エラー"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def generate_draft_sheet(
    column_schema: list[dict],
    dropdown_schema: dict,
    schema_hash: str,
) -> bytes:
    """Generate an empty draft Excel sheet for product registration.

    Returns the workbook as bytes (.xlsx format).
    """
    wb = openpyxl.Workbook()

    # --- Sheet 1: 基本情報 (vertical key-value) ---
    ws1 = wb.active
    ws1.title = "基本情報"
    basic_fields = _resolve_basic_info_fields(column_schema)

    # Headers
    for col_idx, header in enumerate(["項目名", "値（記入欄）", "技術名"], 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, (display, tech_name) in enumerate(basic_fields, 2):
        # A: display name
        cell_a = ws1.cell(row=row_idx, column=1, value=display)
        cell_a.fill = LABEL_FILL
        cell_a.font = LABEL_FONT
        cell_a.border = THIN_BORDER
        # B: value (empty, user fills in)
        cell_b = ws1.cell(row=row_idx, column=2)
        cell_b.border = THIN_BORDER
        # C: technical name (hidden)
        cell_c = ws1.cell(row=row_idx, column=3, value=tech_name)
        cell_c.border = THIN_BORDER

        # Add dropdown if available
        if tech_name in dropdown_schema:
            _add_dropdown_validation(ws1, f"B{row_idx}", dropdown_schema[tech_name])

    # Hide column C
    ws1.column_dimensions["C"].hidden = True
    # Set column widths
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 50

    # --- Sheet 2: 全項目 (horizontal table, all 259 columns) ---
    ws2 = wb.create_sheet("全項目")

    for col_idx, col_def in enumerate(column_schema, 1):
        col_letter = get_column_letter(col_idx)
        display = col_def.get("display_name") or col_def["technical_name"]
        tech_name = col_def["technical_name"]
        sample = col_def.get("sample_value") or ""

        # Row 1: display name header
        cell_h = ws2.cell(row=1, column=col_idx, value=display)
        cell_h.fill = HEADER_FILL
        cell_h.font = HEADER_FONT
        cell_h.border = THIN_BORDER
        cell_h.alignment = Alignment(horizontal="center", wrap_text=True)

        # Row 2: technical name (hidden)
        cell_t = ws2.cell(row=2, column=col_idx, value=tech_name)
        cell_t.border = THIN_BORDER

        # Row 3: sample value
        cell_s = ws2.cell(row=3, column=col_idx, value=sample)
        cell_s.fill = SAMPLE_FILL
        cell_s.font = SAMPLE_FONT
        cell_s.border = THIN_BORDER

        # Row 4: user entry row (empty)
        ws2.cell(row=4, column=col_idx).border = THIN_BORDER

        # Set column width
        ws2.column_dimensions[col_letter].width = max(12, min(30, len(display) * 1.5 + 2))

        # Add dropdown validation for row 4
        if tech_name in dropdown_schema:
            _add_dropdown_validation(ws2, f"{col_letter}4", dropdown_schema[tech_name])

    # Hide row 2 (technical names)
    ws2.row_dimensions[2].hidden = True

    # --- Sheet 3: バリエーション (child variation rows) ---
    ws3 = wb.create_sheet("バリエーション")
    variation_fields = _resolve_variation_fields(column_schema)

    for col_idx, (display, tech_name) in enumerate(variation_fields, 1):
        col_letter = get_column_letter(col_idx)

        # Row 1: display name header
        cell_h = ws3.cell(row=1, column=col_idx, value=display)
        cell_h.fill = HEADER_FILL
        cell_h.font = HEADER_FONT
        cell_h.border = THIN_BORDER
        cell_h.alignment = Alignment(horizontal="center")

        # Row 2: technical name (hidden)
        cell_t = ws3.cell(row=2, column=col_idx, value=tech_name)
        cell_t.border = THIN_BORDER

        ws3.column_dimensions[col_letter].width = max(12, min(30, len(display) * 2 + 2))

        # Add dropdown validation for rows 3-102 (up to 100 variations)
        if tech_name in dropdown_schema:
            _add_dropdown_validation(ws3, f"{col_letter}3:{col_letter}102", dropdown_schema[tech_name])

    # Hide row 2
    ws3.row_dimensions[2].hidden = True

    # --- Sheet 4: _metadata (hidden) ---
    ws4 = wb.create_sheet("_metadata")
    ws4.cell(row=1, column=1, value="draft_sheet_v1")
    ws4.cell(row=2, column=1, value=schema_hash)
    ws4.cell(row=3, column=1, value=datetime.utcnow().isoformat())
    ws4.sheet_state = "hidden"

    # Write to bytes
    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    return output.read()
