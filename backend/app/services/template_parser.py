"""
Parse SIGNAGE.xlsm template to extract schema information.
Extracts: column definitions, validation rules, dropdown values, named ranges.
"""
import openpyxl
from io import BytesIO
import json


def parse_template(file_content: bytes) -> dict:
    """Parse an xlsm/xlsx template file and extract all schema information."""
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=False, keep_vba=True)

    column_schema = _extract_column_schema(wb)
    validation_schema = _extract_validation_schema(wb)
    dropdown_schema = _extract_dropdown_schema(wb)
    defined_names_schema = _extract_defined_names(wb)

    wb.close()

    return {
        "column_schema": column_schema,
        "validation_schema": validation_schema,
        "dropdown_schema": dropdown_schema,
        "defined_names_schema": defined_names_schema,
    }


def _extract_column_schema(wb) -> list[dict]:
    """Extract column definitions from the template sheet rows 3-6."""
    ws = wb["テンプレート"]
    columns = []

    for col_idx in range(1, ws.max_column + 1):
        group_name = ws.cell(row=3, column=col_idx).value
        display_name = ws.cell(row=4, column=col_idx).value
        technical_name = ws.cell(row=5, column=col_idx).value
        sample_value = ws.cell(row=6, column=col_idx).value

        if technical_name is None:
            continue

        columns.append({
            "col_index": col_idx,
            "group_name": str(group_name) if group_name else None,
            "display_name": str(display_name) if display_name else None,
            "technical_name": str(technical_name),
            "sample_value": str(sample_value) if sample_value else None,
        })

    return columns


def _extract_validation_schema(wb) -> list[dict]:
    """Extract data validation rules from the template sheet."""
    ws = wb["テンプレート"]
    validations = []

    if ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            validations.append({
                "ranges": str(dv.sqref),
                "type": dv.type,
                "formula1": str(dv.formula1) if dv.formula1 else None,
                "formula2": str(dv.formula2) if dv.formula2 else None,
                "allow_blank": dv.allow_blank,
                "error_message": dv.error if dv.error else None,
            })

    return validations


def _extract_dropdown_schema(wb) -> dict:
    """Extract dropdown values using named ranges that reference dropdown sheets."""
    dropdowns = {}

    # Use defined names to find dropdown values - these map technical names to cell ranges
    for dn in wb.defined_names.values():
        try:
            for sheet_title, cell_range in dn.destinations:
                if sheet_title in ("Dropdown Lists", "推奨値"):
                    ws = wb[sheet_title]
                    values = []
                    for cell_row in ws[cell_range]:
                        if isinstance(cell_row, tuple):
                            for cell in cell_row:
                                if cell.value is not None:
                                    values.append(str(cell.value))
                        else:
                            if cell_row.value is not None:
                                values.append(str(cell_row.value))
                    if values:
                        dropdowns[dn.name] = values
        except Exception:
            continue

    return dropdowns


def _extract_defined_names(wb) -> list[dict]:
    """Extract all defined names (named ranges) from the workbook."""
    defined_names = []

    for dn in wb.defined_names.values():
        destinations = []
        try:
            for sheet_title, cell_range in dn.destinations:
                destinations.append({
                    "sheet": sheet_title,
                    "range": cell_range,
                })
        except Exception:
            pass

        defined_names.append({
            "name": dn.name,
            "value": str(dn.value) if dn.value else None,
            "destinations": destinations,
        })

    return defined_names
