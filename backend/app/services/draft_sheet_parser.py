"""
Parse a filled-in draft Excel sheet and extract product data.
Returns a dict suitable for creating a Product + Variations.
"""
import openpyxl
from io import BytesIO


def parse_draft_sheet(file_content: bytes, column_schema: list[dict], schema_hash: str) -> dict:
    """Parse an uploaded draft sheet.

    Returns:
        {
            "name": str | None,
            "parent_sku": str | None,
            "variation_theme": str | None,
            "parent_data": dict,
            "variations": [{"child_sku": str|None, "child_data": dict}, ...],
            "warnings": [str, ...],
        }
    """
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    warnings: list[str] = []

    # Build set of known technical names for validation
    known_tech_names = {col["technical_name"] for col in column_schema}

    # --- Validate structure ---
    sheet_names = wb.sheetnames
    if "基本情報" not in sheet_names:
        wb.close()
        raise ValueError("シート「基本情報」が見つかりません。無効な下書きシートです。")

    # --- Check metadata ---
    if "_metadata" in sheet_names:
        ws_meta = wb["_metadata"]
        version = ws_meta.cell(row=1, column=1).value
        if version != "draft_sheet_v1":
            warnings.append("下書きシートのバージョンが不明です")
        file_hash = ws_meta.cell(row=2, column=1).value
        if file_hash and file_hash != schema_hash:
            warnings.append("テンプレートバージョンが異なります。一部の項目が正しくマッピングされない可能性があります。")

    # --- Sheet 1: 基本情報 ---
    ws1 = wb["基本情報"]
    name = None
    parent_sku = None
    variation_theme = None
    parent_data: dict = {}

    for row_idx in range(2, ws1.max_row + 1):
        tech_name = ws1.cell(row=row_idx, column=3).value  # Column C
        value = ws1.cell(row=row_idx, column=2).value  # Column B
        if not tech_name or value is None or str(value).strip() == "":
            continue
        value = str(value).strip()
        tech_name = str(tech_name).strip()

        if tech_name == "_app_name":
            name = value
        elif tech_name == "_variation_theme":
            variation_theme = value
        elif "contribution_sku" in tech_name:
            parent_sku = value
            parent_data[tech_name] = value
        else:
            if tech_name not in known_tech_names and not tech_name.startswith("_"):
                warnings.append(f"基本情報: 不明な技術名「{tech_name}」をスキップしました")
                continue
            parent_data[tech_name] = value

    # --- Sheet 2: 全項目 ---
    if "全項目" in sheet_names:
        ws2 = wb["全項目"]
        # Row 2: technical names
        tech_names_row = {}
        for col_idx in range(1, ws2.max_column + 1):
            tn = ws2.cell(row=2, column=col_idx).value
            if tn:
                tech_names_row[col_idx] = str(tn).strip()

        if not tech_names_row:
            warnings.append("全項目シートの技術名行が空です")
        else:
            # Row 4: user values
            for col_idx, tech_name in tech_names_row.items():
                value = ws2.cell(row=4, column=col_idx).value
                if value is None or str(value).strip() == "":
                    continue
                value = str(value).strip()
                # Basic info sheet values take priority (don't overwrite)
                if tech_name not in parent_data:
                    if tech_name not in known_tech_names:
                        warnings.append(f"全項目: 不明な技術名「{tech_name}」をスキップしました")
                        continue
                    parent_data[tech_name] = value

    # --- Sheet 3: バリエーション ---
    variations: list[dict] = []
    if "バリエーション" in sheet_names:
        ws3 = wb["バリエーション"]
        # Row 2: technical names
        var_tech_names = {}
        for col_idx in range(1, ws3.max_column + 1):
            tn = ws3.cell(row=2, column=col_idx).value
            if tn:
                var_tech_names[col_idx] = str(tn).strip()

        # Row 3+: variation data
        for row_idx in range(3, ws3.max_row + 1):
            child_data = {}
            child_sku = None
            has_data = False

            for col_idx, tech_name in var_tech_names.items():
                value = ws3.cell(row=row_idx, column=col_idx).value
                if value is None or str(value).strip() == "":
                    continue
                value = str(value).strip()
                has_data = True

                if "contribution_sku" in tech_name:
                    child_sku = value
                child_data[tech_name] = value

            if has_data:
                variations.append({
                    "child_sku": child_sku,
                    "child_data": child_data,
                })

    wb.close()

    return {
        "name": name,
        "parent_sku": parent_sku,
        "variation_theme": variation_theme,
        "parent_data": parent_data,
        "variations": variations,
        "warnings": warnings,
    }
